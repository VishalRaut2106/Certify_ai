from fastapi import FastAPI, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from typing import List
from dotenv import load_dotenv
load_dotenv()  # Load .env from backend/ directory
from ocr import extract_text
from qr_decoder import decode_qr
from comparator import compare_fields
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import os
import sys
import asyncio
import uuid
import io
from concurrent.futures import ThreadPoolExecutor

app = FastAPI()

# ── Resolve frontend path (dev vs frozen .exe) ─────────────────────────────
if getattr(sys, 'frozen', False):
    FRONTEND_DIR = os.path.join(sys._MEIPASS, 'frontend')
else:
    FRONTEND_DIR = os.path.join(os.path.dirname(__file__), '..', 'frontend')

app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

# Optimize for HF Spaces 2 vCPUs: limit Tesseract threads and run 4 parallel workers
os.environ["OMP_THREAD_LIMIT"] = "1"
executor = ThreadPoolExecutor(max_workers=4)




@app.get("/")
def index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


@app.post("/verify")
async def verify_certificate(certificate: UploadFile = File(...)):
    ext = os.path.splitext(certificate.filename)[1]
    temp_path = f"temp_{uuid.uuid4().hex}{ext}"
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(certificate.file, buffer)

    try:
        extracted = extract_text(temp_path)
        qr_data   = decode_qr(temp_path)

        if not qr_data:
            return {
                "name":    {"ocr": extracted.get("name"),   "qr": None, "match": False},
                "course":  {"ocr": extracted.get("course"), "qr": None, "match": False},
                "date":    {"ocr": extracted.get("date"),   "qr": None, "match": False},
                "verdict": "Manual Review - QR Unreadable"
            }

        return compare_fields(extracted, qr_data)

    except Exception as e:
        import traceback; traceback.print_exc()
        return {
            "name":    {"ocr": None, "qr": None, "match": False},
            "course":  {"ocr": None, "qr": None, "match": False},
            "date":    {"ocr": None, "qr": None, "match": False},
            "verdict": "Manual Review - Processing Error"
        }
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def process_single(filename, temp_path):
    """Runs in thread pool — handles OCR + QR + compare for one file."""
    try:
        extracted = extract_text(temp_path)
        qr_data   = decode_qr(temp_path)

        if not qr_data:
            return {
                'filename':  filename,
                'name':      extracted.get('name',   '—'),
                'issued_by': 'Unknown',
                'course':    extracted.get('course', '—'),
                'date':      extracted.get('date',   '—'),
                'flag':      'Manual Review - QR Unreadable'
            }

        comparison = compare_fields(extracted, qr_data)
        return {
            'filename':  filename,
            'name':      extracted.get('name') or '—',  # Name from certificate (OCR) - could be edited
            'issued_by': qr_data.get('name') or 'Unknown',  # Original person's name from QR (issuedTo)
            'course':    qr_data.get('course') or extracted.get('course') or '—',
            'date':      qr_data.get('date') or extracted.get('date') or '—',
            'flag':      comparison['verdict']
        }

    except Exception as e:
        print(f'Error processing {filename}: {e}')
        return {
            'filename':  filename,
            'name':      '—',
            'issued_by': 'Unknown',
            'course':    '—',
            'date':      '—',
            'flag':      'Manual Review - Processing Error'
        }
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def build_excel(results) -> bytes:
    """Build the Excel workbook and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verification Results"

    # Professional badge blue for headers
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    # Thin borders for all cells
    thin_border = Border(left=Side(style='thin', color='E5E7EB'), 
                         right=Side(style='thin', color='E5E7EB'), 
                         top=Side(style='thin', color='E5E7EB'), 
                         bottom=Side(style='thin', color='E5E7EB'))

    headers = ["File Name", "Name on Certificate", "Issued By", "Course", "Date", "Result"]
    for col, header in enumerate(headers, 2):  # Start at column 2 (B)
        cell = ws.cell(row=2, column=col, value=header)  # Start at row 2
        cell.font  = header_font
        cell.fill  = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Increase header row height
    ws.row_dimensions[2].height = 25  # Changed from row 1 to row 2

    # Status badge colors
    green_fill  = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Light green
    red_fill    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light red
    orange_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light yellow/orange

    for row, result in enumerate(results, 3):  # Start data at row 3
        ws.cell(row=row, column=2, value=result['filename']).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=3, value=result['name']).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=4, value=result['issued_by']).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=5, value=result['course']).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=6, value=result['date']).alignment = Alignment(horizontal="center", vertical="center")
        
        # Result column centered
        ws.cell(row=row, column=7, value=result['flag']).alignment = Alignment(horizontal="center", vertical="center")

        flag = result['flag']
        fill = green_fill if flag == 'Verified' else (orange_fill if 'Manual' in flag else red_fill)

        for col in range(2, 8):  # Columns 2 to 7
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            
        # Give data rows a little padding
        ws.row_dimensions[row].height = 20

    # Adjust column widths for the offset
    ws.column_dimensions['A'].width = 3   # Small margin for column A
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30

    # Calculate summary counts
    verified_cnt = 0
    manual_cnt = 0
    fraud_cnt = 0
    for res in results:
        flag = res.get('flag', '')
        if flag == 'Verified':
            verified_cnt += 1
        elif 'Manual' in flag:
            manual_cnt += 1
        else:
            fraud_cnt += 1

    # Add summary section at the bottom
    summary_start_row = len(results) + 4
    
    # Summary Header
    summary_header_cell = ws.cell(row=summary_start_row, column=2, value="Verification Summary")
    summary_header_cell.font = Font(bold=True, size=11, color="FFFFFF")
    summary_header_cell.fill = header_fill
    summary_header_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Merge cells for header
    ws.merge_cells(start_row=summary_start_row, start_column=2, end_row=summary_start_row, end_column=3)

    summary_data = [
        ("Total Verified", verified_cnt, green_fill),
        ("Needs Manual Review", manual_cnt, orange_fill),
        ("Possible Fraud", fraud_cnt, red_fill),
        ("Total Processed", len(results), PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"))
    ]

    for i, (label, count, fill) in enumerate(summary_data):
        curr_row = summary_start_row + 1 + i
        label_cell = ws.cell(row=curr_row, column=2, value=label)
        count_cell = ws.cell(row=curr_row, column=3, value=count)
        
        label_cell.font = Font(bold=True)
        label_cell.fill = fill
        label_cell.border = thin_border
        
        count_cell.font = Font(bold=True)
        count_cell.fill = fill
        count_cell.border = thin_border
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[curr_row].height = 20

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


@app.post("/verify-bulk")
async def verify_bulk(
    certificates: List[UploadFile] = File(...)
):
    # Save all files with UUID names (avoids space issues)
    tasks = []
    for cert in certificates:
        ext       = os.path.splitext(cert.filename)[1]
        temp_path = f"temp_{uuid.uuid4().hex}{ext}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(cert.file, buffer)
        tasks.append((cert.filename, temp_path))

    # Process all in parallel
    loop    = asyncio.get_event_loop()
    futures = [
        loop.run_in_executor(executor, process_single, filename, temp_path)
        for filename, temp_path in tasks
    ]
    results = await asyncio.gather(*futures)

    # Build Excel and stream as download
    excel_bytes = build_excel(results)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="verification_results.xlsx"'}
    )


from pydantic import BaseModel
from typing import Optional, Dict, Any

class VerificationResult(BaseModel):
    filename: str
    name: Optional[Dict[str, Any]] = None
    issued_by: Optional[Dict[str, Any]] = None
    course: Optional[Dict[str, Any]] = None
    date: Optional[Dict[str, Any]] = None
    verdict: str

@app.post("/generate-excel")
async def generate_excel(results_payload: List[Dict[str, Any]]):
    """
    Accepts a list of verification results from the frontend and returns an Excel file.
    This enables real-time progress tracking on the client side.
    """
    # Transform the payload to match what build_excel expects
    # build_excel expects a list of dicts with:
    # 'filename', 'name', 'issued_by', 'course', 'date', 'flag'
    
    formatted_results = []
    for res in results_payload:
        formatted_results.append({
            'filename': res.get('filename', 'Unknown'),
            'name': res.get('name', {}).get('qr', '—') if res.get('name') else '—',
            'issued_by': res.get('issued_by', {}).get('qr', '—') if res.get('issued_by') else '—',
            'course': res.get('course', {}).get('qr', '—') if res.get('course') else '—',
            'date': res.get('date', {}).get('qr', '—') if res.get('date') else '—',
            'flag': res.get('verdict', 'Error')
        })

    excel_bytes = build_excel(formatted_results)
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="verification_results.xlsx"'}
    )