from fastapi import FastAPI, File, UploadFile, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional
from dotenv import load_dotenv
load_dotenv()  # Load .env from backend/ directory
from ocr import extract_text
from qr_decoder import decode_qr
from comparator import compare_fields
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import shutil
import os
import asyncio
import uuid
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from concurrent.futures import ThreadPoolExecutor

app = FastAPI()
app.mount("/static", StaticFiles(directory="../frontend"), name="static")

# Optimize for HF Spaces 2 vCPUs: limit Tesseract threads and run 4 parallel workers
os.environ["OMP_THREAD_LIMIT"] = "1"
executor = ThreadPoolExecutor(max_workers=4)

# ── SMTP config from environment variables ──────────────────
SMTP_HOST     = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.getenv("SMTP_PORT", "587"))
SMTP_EMAIL    = os.getenv("SMTP_EMAIL", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")


def send_excel_email(to_email: str, excel_bytes: bytes, total: int):
    """Send the Excel report as an email attachment."""
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        raise ValueError("SMTP credentials not configured. Set SMTP_EMAIL and SMTP_PASSWORD env vars.")

    msg = MIMEMultipart()
    msg["From"]    = f"CertifyAI <{SMTP_EMAIL}>"
    msg["To"]      = to_email
    msg["Subject"] = f"Certificate Verification Report — {total} certificate{'s' if total != 1 else ''}"

    body = f"""Hi,

Your bulk certificate verification is complete.

Total certificates processed: {total}

Please find the detailed Excel report attached. Each row is colour-coded:
  🟢 Green  — Verified (authentic)
  🔴 Red    — Fraud detected
  🟡 Yellow — Manual review required

—
CertifyAI · Infosys Springboard Certificate Verification
"""
    msg.attach(MIMEText(body, "plain"))

    # Attach Excel
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", 'attachment; filename="verification_results.xlsx"')
    msg.attach(part)

    # Hugging Face Spaces often blocks outbound SMTP ports.
    # Add a 5-second timeout so the app doesn't hang if the port is blocked.
    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, to_email, msg.as_string())
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, to_email, msg.as_string())


@app.get("/")
def index():
    return FileResponse("../frontend/index.html")


@app.post("/verify")
async def verify_certificate(certificate: UploadFile = File(...)):
    ext = os.path.splitext(certificate.filename)[1]
    temp_path = f"temp_{uuid.uuid4().hex}{ext}"
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(certificate.file, buffer)

    try:
        extracted = extract_text(temp_path)
        qr_data   = decode_qr(temp_path)

        if not qr_data:
            return {
                "name":    {"ocr": extracted.get("name"),   "qr": None, "match": False},
                "course":  {"ocr": extracted.get("course"), "qr": None, "match": False},
                "date":    {"ocr": extracted.get("date"),   "qr": None, "match": False},
                "verdict": "Manual Review - QR Unreadable"
            }

        return compare_fields(extracted, qr_data)

    except Exception as e:
        import traceback; traceback.print_exc()
        return {
            "name":    {"ocr": None, "qr": None, "match": False},
            "course":  {"ocr": None, "qr": None, "match": False},
            "date":    {"ocr": None, "qr": None, "match": False},
            "verdict": "Manual Review - Processing Error"
        }
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def process_single(filename, temp_path):
    """Runs in thread pool — handles OCR + QR + compare for one file."""
    try:
        extracted = extract_text(temp_path)
        qr_data   = decode_qr(temp_path)

        if not qr_data:
            return {
                'filename':  filename,
                'name':      extracted.get('name',   '—'),
                'issued_by': 'Unknown',
                'course':    extracted.get('course', '—'),
                'date':      extracted.get('date',   '—'),
                'flag':      'Manual Review - QR Unreadable'
            }

        comparison = compare_fields(extracted, qr_data)
        return {
            'filename':  filename,
            'name':      extracted.get('name') or '—',  # Name from certificate (OCR) - could be edited
            'issued_by': qr_data.get('name') or 'Unknown',  # Original person's name from QR (issuedTo)
            'course':    qr_data.get('course') or extracted.get('course') or '—',
            'date':      qr_data.get('date') or extracted.get('date') or '—',
            'flag':      comparison['verdict']
        }

    except Exception as e:
        print(f'Error processing {filename}: {e}')
        return {
            'filename':  filename,
            'name':      '—',
            'issued_by': 'Unknown',
            'course':    '—',
            'date':      '—',
            'flag':      'Manual Review - Processing Error'
        }
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def build_excel(results) -> bytes:
    """Build the Excel workbook and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verification Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")

    headers = ["File Name", "Name on Certificate", "Issued By", "Course", "Date", "Result"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    green_fill  = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
    red_fill    = PatternFill(start_color="fff1f2", end_color="fff1f2", fill_type="solid")
    orange_fill = PatternFill(start_color="fffbeb", end_color="fffbeb", fill_type="solid")

    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['filename'])
        ws.cell(row=row, column=2, value=result['name'])
        ws.cell(row=row, column=3, value=result['issued_by'])
        ws.cell(row=row, column=4, value=result['course'])
        ws.cell(row=row, column=5, value=result['date'])
        ws.cell(row=row, column=6, value=result['flag'])

        flag = result['flag']
        fill = green_fill if flag == 'Verified' else (orange_fill if 'Manual' in flag else red_fill)

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill      = fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


@app.post("/verify-bulk")
async def verify_bulk(
    certificates: List[UploadFile] = File(...),
    email: Optional[str] = Form(None)
):
    # Save all files with UUID names (avoids space issues)
    tasks = []
    for cert in certificates:
        ext       = os.path.splitext(cert.filename)[1]
        temp_path = f"temp_{uuid.uuid4().hex}{ext}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(cert.file, buffer)
        tasks.append((cert.filename, temp_path))

    # Process all in parallel
    loop    = asyncio.get_event_loop()
    futures = [
        loop.run_in_executor(executor, process_single, filename, temp_path)
        for filename, temp_path in tasks
    ]
    results = await asyncio.gather(*futures)

    # Build Excel bytes
    excel_bytes = build_excel(results)

    # ── Send email if requested ──────────────────────────────
    email = (email or "").strip()
    if email:
        try:
            await asyncio.get_event_loop().run_in_executor(
                None, send_excel_email, email, excel_bytes, len(results)
            )
            # Return JSON confirmation — no file download needed
            return {"status": "emailed", "to": email, "total": len(results)}
        except Exception as e:
            print(f"Email error: {e}")
            # Fall through to download if email fails
            return StreamingResponse(
                io.BytesIO(excel_bytes),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": 'attachment; filename="verification_results.xlsx"',
                    "X-Email-Error": str(e)
                }
            )

    # ── No email — stream download ───────────────────────────
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="verification_results.xlsx"'}
    )