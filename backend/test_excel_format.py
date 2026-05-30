"""
Test script to verify the Excel report format includes the new "Issued By" column
"""
import sys
import os

# Add backend directory to path
sys.path.insert(0, os.path.dirname(__file__))

from app import build_excel
import openpyxl
import io

# Create sample test data
test_results = [
    {
        'filename': 'test_certificate_1.pdf',
        'name': 'John Doe',  # Name on certificate (OCR - could be edited)
        'issued_by': 'John Doe',  # Original name from QR (authentic)
        'course': 'Python Programming',
        'date': 'January 15, 2024',
        'flag': 'Verified'
    },
    {
        'filename': 'test_certificate_2.pdf',
        'name': 'FAKE NAME',  # Edited name on certificate (OCR)
        'issued_by': 'Jane Smith',  # Original authentic name from QR
        'course': 'Machine Learning',
        'date': 'February 20, 2024',
        'flag': 'Fraud Detected'
    },
    {
        'filename': 'test_certificate_3.pdf',
        'name': 'Bob Johnson',  # Name on certificate (OCR)
        'issued_by': 'Unknown',  # QR unreadable
        'course': 'Data Science',
        'date': 'March 10, 2024',
        'flag': 'Manual Review - QR Unreadable'
    }
]

# Build Excel
print("Building Excel report with test data...")
excel_bytes = build_excel(test_results)

# Load and verify the Excel structure
wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
ws = wb.active

print("\n✅ Excel Report Generated Successfully!")
print("\n" + "="*80)
print("EXCEL REPORT STRUCTURE")
print("="*80)

# Print headers
headers = []
for col in range(1, 7):
    header = ws.cell(row=1, column=col).value
    headers.append(header)
    print(f"Column {col}: {header}")

print("\n" + "="*80)
print("SAMPLE DATA ROWS")
print("="*80)

# Print data rows
for row in range(2, 5):
    print(f"\nRow {row-1}:")
    for col in range(1, 7):
        cell_value = ws.cell(row=row, column=col).value
        print(f"  {headers[col-1]}: {cell_value}")

print("\n" + "="*80)
print("VERIFICATION")
print("="*80)

# Verify the structure
expected_headers = ["File Name", "Name on Certificate", "Issued By", "Course", "Date", "Result"]
if headers == expected_headers:
    print("✅ Headers match expected format!")
    print(f"   Expected: {expected_headers}")
    print(f"   Got:      {headers}")
else:
    print("❌ Headers do NOT match!")
    print(f"   Expected: {expected_headers}")
    print(f"   Got:      {headers}")

# Check if "Issued By" column exists
if "Issued By" in headers:
    print("✅ 'Issued By' column is present!")
    issued_by_col = headers.index("Issued By") + 1
    print(f"   Located at column {issued_by_col}")
    
    # Show issued by values
    print("\n   Issued By values in test data:")
    for row in range(2, 5):
        issued_by_value = ws.cell(row=row, column=issued_by_col).value
        name_value = ws.cell(row=row, column=2).value
        print(f"   - {name_value}: Issued by '{issued_by_value}'")
else:
    print("❌ 'Issued By' column is MISSING!")

print("\n" + "="*80)
print("TEST COMPLETE")
print("="*80)
